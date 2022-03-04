import openpyxl

class excel_file():
    def __init__(self):
        self.wb, self.ws = self.getfile()

    def run(self):
        pass

    def getfile(self):
        try:
            wb = openpyxl.load_workbook("收敛情况.xlsx")
            sheet = wb["收敛情况"]
        except:
            wb = openpyxl.Workbook()
            wb.create_sheet(index=0, title="收敛情况")
            sheet = wb['收敛情况']
            sheet["A1"] = "时间"
            sheet["B1"] = "车流量"
            sheet["C1"] = "收敛值"
            sheet["D1"] = "最大值"
            wb.save('收敛情况.xlsx')
        return wb,sheet

    def write(self,data):
        self.ws.append(data)

    def save(self):
        self.wb.save('收敛情况.xlsx')

    def close_with_save(self):
        self.save()
        self.wb.close()


class excel_file2():
    def __init__(self):
        self.wb, self.ws = self.getfile()

    def run(self):
        pass

    def getfile(self):
        try:
            wb = openpyxl.load_workbook("冗余情况.xlsx")
            sheet = wb["冗余情况"]
        except:
            wb = openpyxl.Workbook()
            wb.create_sheet(index=0, title="冗余情况")
            sheet = wb['冗余情况']
            sheet["A1"] = "时间"
            sheet["B1"] = "GPSR链路"
            sheet["C1"] = "深度强化学习链路"
            wb.save('冗余情况.xlsx')
        return wb,sheet

    def write(self,data):
        self.ws.append(data)

    def save(self):
        self.wb.save('冗余情况.xlsx')

    def close_with_save(self):
        self.save()
        self.wb.close()